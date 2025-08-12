from __future__ import annotations
from typing import Dict, List, Optional, Tuple
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.utils.cell import range_boundaries, coordinate_from_string
from openpyxl.worksheet.datavalidation import DataValidation


DEFAULT_DATA_SHEET = "Fichas 2025"   
DEFAULT_HEADER_ROW = 2                


TABLES: Dict[str, Tuple[str, str]] = {
    # clave → (nombre_tabla_excel, cabecera_exacta)
    "USUARIOS_HACE_FICHA": ("HACE_FICHA", "TRABAJADORA QUE HACE LA FICHA"),
    "USUARIOS_SUBE_FICHA": ("SUBE_FICHA", "TRABAJADORA QUE SUBE LA FICHA"),
    "PORTALES": ("PORTALES", "PORTALES"),
    "TEMATICAS": ("TEMATICAS", "TEMATICA 1-2-3"),
    "CCAA": ("COMUNIDADES", "CCAA"),
    "PROVINCIAS": ("DIPUTACIONES", "DIPUTACIÓN"),
    "ESTADO_UE": ("ESTADO_EUROPA", "ESTADO/EUROPA"),
}


def _dedup(seq: List[str]) -> List[str]:
    seen = set(); out = []
    for s in seq:
        s = s.strip()
        if s and s not in seen:
            seen.add(s); out.append(s)
    return out


# ------------------------
# 1) Desde Tablas (si existen)
# ------------------------

def _extract_from_table(wb, table_name: str, header_label: str) -> List[str]:
    for ws in wb.worksheets:
        tbl = ws.tables.get(table_name)
        if not tbl:
            continue
        ref = getattr(tbl, "ref", None) or str(tbl)
        min_c, min_r, max_c, max_r = range_boundaries(ref)
        headers = {}
        for c in range(min_c, max_c + 1):
            v = ws.cell(row=min_r, column=c).value
            if isinstance(v, str) and v.strip():
                headers[v.strip().upper()] = c
        # match flexible con acentos y espacios
        def _norm(s: str) -> str:
            t = s.upper()
            rep = str.maketrans("ÁÉÍÓÚÜÑ ", "AEIOUUN_ ")
            return t.translate(rep)
        col = headers.get(header_label.strip().upper())
        if not col:
            for k, cidx in headers.items():
                if _norm(k) == _norm(header_label):
                    col = cidx; break
        if not col:
            return []
        vals: List[str] = []
        for r in range(min_r + 1, max_r + 1):
            v = ws.cell(row=r, column=col).value
            if v is not None and str(v).strip() != "":
                vals.append(str(v).strip())
        return _dedup(vals)
    return []


# ------------------------
# 2) Desde Data Validations (dinámico real)
# ------------------------

def _read_range_values(wb, range_ref: str) -> List[str]:
    # Admite: 'Hoja'!$A$2:$A$40   |   $A$2:$A$40 (siempre requiere hoja)
    if "!" in range_ref:
        sheet_name, rng = range_ref.split("!", 1)
        sheet_name = sheet_name.strip().strip("'")
        rng = rng.strip()
        ws = wb[sheet_name]
    else:
        # sin nombre de hoja no podemos resolver de forma global
        return []
    min_c, min_r, max_c, max_r = range_boundaries(rng)
    vals: List[str] = []
    for r in range(min_r, max_r + 1):
        for c in range(min_c, max_c + 1):
            v = ws.cell(row=r, column=c).value
            if v is not None and str(v).strip() != "":
                vals.append(str(v).strip())
    return _dedup(vals)


def _header_by_cell(ws, header_row: int, coord: str) -> Optional[str]:
    col_letter, _ = coordinate_from_string(coord)
    # traer índice de columna desde letra
    from openpyxl.utils import column_index_from_string
    col_idx = column_index_from_string(col_letter)
    v = ws.cell(row=header_row, column=col_idx).value
    return str(v) if v is not None else None


def _collect_validations(wb, data_sheet: str, header_row: int) -> Dict[str, List[str]]:
    ws = wb[data_sheet]
    out: Dict[str, List[str]] = {}
    dvs = ws.data_validations
    if not dvs:
        return out

    for dv in dvs.dataValidation:  # DataValidation
        if dv.type != "list" or not dv.formula1:
            continue

        f = str(dv.formula1).strip()
        # Quitar '=' inicial típico en fórmulas de validación
        if f.startswith("="):
            f = f[1:].strip()

        values: List[str] = []

        if f.startswith('"') and f.endswith('"'):
            # lista inline: "A,B,C"
            values = [s.strip() for s in f.strip('"').split(",") if s.strip()]
        else:
            # referencia a rango o nombre definido
            if "!" in f or ":" in f:
                # Rango explícito
                values = _read_range_values(wb, f)
            else:
                # Nombre definido (puede devolver objeto o lista de objetos)
                dn_obj = wb.defined_names.get(f)
                dn_list = dn_obj if isinstance(dn_obj, list) else [dn_obj] if dn_obj else []
                for dn in dn_list:
                    if dn is None:
                        continue
                    try:
                        dests = list(dn.destinations)  # [(sheetname, ref), ...]
                    except Exception:
                        # Nombre definido no-resoluble; lo ignoramos
                        continue
                    for sheetname, ref in dests:
                        if not sheetname or not ref:
                            continue
                        rng_ref = f"'{sheetname}'!{ref}"
                        vals = _read_range_values(wb, rng_ref)
                        if vals:
                            values = _dedup(values + vals)

        if not values:
            continue

        # Mapear la validación a la cabecera de su(s) columna(s)
        for cell_range in dv.ranges:
            min_c, _, _, _ = range_boundaries(str(cell_range))
            from openpyxl.utils import get_column_letter
            col_letter = get_column_letter(min_c)
            header_cell = f"{col_letter}{header_row}"
            header = _header_by_cell(ws, header_row, header_cell)
            if header:
                out.setdefault(header, [])
                out[header] = _dedup(out[header] + values)

    return out



# ------------------------
# 3) API pública
# ------------------------

def load_enums_from_bytes(
    excel_bytes: bytes,
    data_sheet: str = DEFAULT_DATA_SHEET,
    header_row: int = DEFAULT_HEADER_ROW,
) -> Dict[str, List[str]]:
    """Carga enums de forma dinámica.
    1) Si existen Tablas con los nombres de TABLES, se añaden.
    2) Se leen también los Data Validations (listas) de la hoja de datos indicada.
    El resultado es un dict con claves de cabecera EXACTAS tal como aparecen en la hoja de datos,
    más las claves de TABLES (si existen), sin duplicados.
    """
    wb = load_workbook(BytesIO(excel_bytes), data_only=True)

    enums: Dict[str, List[str]] = {}

    # A) Data validations (dinámico y preferente para UI)
    try:
        by_header = _collect_validations(wb, data_sheet, header_row)
        enums.update(by_header)
    except KeyError:
        # la hoja no existe; ignora esta parte
        pass

    # B) Tablas estructuradas (si las hubiera)
    for key, (tbl, col) in TABLES.items():
        vals = _extract_from_table(wb, tbl, col)
        if vals:
            enums[key] = vals

    return enums

from typing import Dict, Any, List
from io import BytesIO
from openpyxl import load_workbook
import unicodedata
import logging
import re

# Configuración de la hoja
HEADER_ROW = 2       # cabeceras en fila 2
DATA_START_ROW = 3   # datos empiezan en fila 3
DEFAULT_SHEET = "Fichas 2025"

# Columnas especiales (si existen en el Excel)
AMBITO_COLS = [
    "AMBITO UE/ESTADO",
    "AMBITO CC AA",
    "AMBITO PROVINCIAL",
    "AMBITO MUNICIPAL",
]
PORTAL_COLS = ["Mayores", "Discapacidad", "Familia", "Mujer", "Salud"]
TEMATICA_COLS = ["TEMÁTICA 1", "TEMÁTICA 2", "TEMÁTICA 3"]

# Estas columnas se usarán por defecto para decidir si una fila está "ocupada".
# Puedes cambiarlas si tu plantilla se apoya en otras celdas clave.
DEFAULT_REQUIRED_COLS = ["NOMBRE DE FICHA", "VENCIMIENTO"]

logger = logging.getLogger(__name__)


def norm_header(s: str) -> str:
    s = s.strip()
    s = re.sub(r"\s+", " ", s)
    s = "".join(
        c for c in unicodedata.normalize("NFD", s)
        if unicodedata.category(c) != "Mn"
    )
    return s.upper()


def _norm(s: str) -> str:
    return (s or "").strip().upper()


def _headers_index(ws) -> Dict[str, int]:
    """Devuelve un mapa normalizado nombre_de_columna -> índice (1-based)."""
    idx = {}
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=HEADER_ROW, column=c).value
        if v and str(v).strip():
            idx[_norm(str(v))] = c
    logger.info("Cabeceras detectadas (%d): %s", len(idx), list(idx.keys())[:10])
    return idx


def _first_empty_row(
    ws,
    headers: Dict[str, int],
    required_cols: List[str] | None = None,
) -> int:
    """
    Devuelve la primera fila "vacía de verdad" mirando únicamente columnas relevantes.
    Una fila se considera vacía si TODAS las columnas relevantes están vacías (None o "").
    """
    if not required_cols:
        required_cols = DEFAULT_REQUIRED_COLS

    # Nos quedamos sólo con los headers que existen realmente en la hoja
    required_indices = [headers[_norm(h)] for h in required_cols if _norm(h) in headers]
    if not required_indices:
        # Fallback: usa todas las columnas conocidas del header
        required_indices = list(headers.values())

    r = DATA_START_ROW
    while True:
        is_empty = True
        for c in required_indices:
            v = ws.cell(row=r, column=c).value
            if v not in (None, ""):
                is_empty = False
                break
        if is_empty:
            logger.info("Primera fila vacía detectada: %s", r)
            return r
        r += 1


def _set_if(header: str, value, ws, row: int, headers: Dict[str, int]):
    col = headers.get(_norm(header))
    if col:
        ws.cell(row=row, column=col).value = value
        logger.debug("Escrito [%s] en fila %d, col %d: %r", header, row, col, value)
    else:
        logger.debug("Header no encontrado, NO se escribe: %s", header)


def _clear(headers, ws, row: int, cols: list[str]):
    for h in cols:
        _set_if(h, "", ws, row, headers)


def _apply_ambito_exclusive(ws, row: int, headers: Dict[str, int], payload: Dict[str, Any]):
    """Solo un ÁMBITO puede quedar informado a la vez.
    Si en el payload viene alguno con valor, se limpian los demás y se escribe ese.
    """
    has_any = any(payload.get(k) for k in AMBITO_COLS)
    if not has_any:
        return
    for col in AMBITO_COLS:
        if payload.get(col):
            _clear(headers, ws, row, AMBITO_COLS)
            _set_if(col, payload[col], ws, row, headers)
            logger.info("Ámbito exclusivo aplicado: %s=%r", col, payload[col])
            return


def write_auto_fields(
    excel_bytes: bytes,
    auto_fields: Dict[str, Any],
    sheet: str = DEFAULT_SHEET,
    required_cols: List[str] | None = None,
) -> Dict[str, Any]:
    """
    Escribe `auto_fields` en la primera fila libre detectada, sin generar ningún ID.
    - Respeta el marco/estilos porque no inserta filas ni columnas.
    - required_cols te permite definir qué columnas marcan que una fila está ocupada.
    """
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb[sheet]
    headers = _headers_index(ws)

    row = _first_empty_row(ws, headers, required_cols)
    logger.info("Escritura en hoja '%s', fila %d (base 1)", ws.title, row)

    # 1) Portales (si existen en auto_fields)
    for col in PORTAL_COLS:
        if col in auto_fields:
            _set_if(col, auto_fields[col], ws, row, headers)

    # 2) Temáticas (1..3)
    for col in TEMATICA_COLS:
        if col in auto_fields:
            _set_if(col, auto_fields[col], ws, row, headers)

    # 3) Ámbito exclusivo (si viene algún ámbito informado)
    _apply_ambito_exclusive(ws, row, headers, auto_fields)

    # 4) Resto de campos
    for k, v in auto_fields.items():
        if k in PORTAL_COLS or k in TEMATICA_COLS or k in AMBITO_COLS:
            continue
        _set_if(k, v, ws, row, headers)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    base0 = row - 1
    logger.info("Guardado. Hoja=%s, fila(base0)=%d", ws.title, base0)

    return {
        "sheet": ws.title,
        "row": base0,  # índice base 0 de la fila escrita
        "updated_excel_bytes": out.getvalue(),
    }


def update_row_in_excel(
    excel_bytes: bytes,
    sheet: str,
    row_index_base0: int,
    updates: Dict[str, Any],
) -> Dict[str, Any]:
    """Actualiza una fila existente (row_index_base0) con los pares clave/valor de `updates`."""
    wb = load_workbook(BytesIO(excel_bytes))
    ws = wb[sheet]
    headers = _headers_index(ws)
    row = row_index_base0 + 1

    logger.info("Actualizar fila (base1) %d en hoja '%s'", row, ws.title)

    for k, v in (updates or {}).items():
        _set_if(k, v, ws, row, headers)

    # Reaplica la exclusividad de ÁMBITO si alguno de ellos se tocó
    if any(k in updates for k in AMBITO_COLS):
        payload = {
            k: ws.cell(row=row, column=headers[_norm(k)]).value if _norm(k) in headers else ""
            for k in AMBITO_COLS
        }
        _apply_ambito_exclusive(ws, row, headers, payload)

    out = BytesIO()
    wb.save(out)
    out.seek(0)

    logger.info("Actualización guardada. Fila(base0)=%d", row_index_base0)
    return {
        "sheet": ws.title,
        "row": row_index_base0,
        "updated_excel_bytes": out.getvalue(),
    }
